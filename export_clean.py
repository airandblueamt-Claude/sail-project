"""Export the equipment inventory to a clean Excel file for team review."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_db
import os

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "SAIL_Equipment_Clean_v2.xlsx")


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipment Inventory"

    # ── Styles ───────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D3142", end_color="2D3142", fill_type="solid")
    cat_font = Font(bold=True, size=11, color="1A1D2E")
    cat_fill = PatternFill(start_color="E2E4EC", end_color="E2E4EC", fill_type="solid")
    bookable_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    flag_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="E2E4EC")
    )
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="top")

    # ── Headers ──────────────────────────────────────────────────────
    headers = ["#", "Category", "Description", "Brand", "Model / Specs",
               "Qty", "Unit", "Bookable", "Review Flag", "Notes"]
    col_widths = [5, 25, 45, 20, 55, 8, 8, 10, 16, 40]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J1"

    # ── Data ─────────────────────────────────────────────────────────
    with get_db() as conn:
        rows = conn.execute("""
            SELECT em.id, c.name as category, em.name as description,
                   em.brand, em.specifications, em.unit, em.expected_qty,
                   em.is_bookable, em.notes
            FROM equipment_models em
            JOIN categories c ON em.category_id = c.id
            ORDER BY c.name, em.name
        """).fetchall()

    current_cat = None
    row_num = 1
    item_num = 0

    for r in rows:
        # Category separator row
        if r['category'] != current_cat:
            current_cat = r['category']
            row_num += 1
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = cat_fill
            ws.cell(row=row_num, column=2, value=current_cat).font = cat_font
            ws.merge_cells(start_row=row_num, start_column=2,
                           end_row=row_num, end_column=len(headers))

        row_num += 1
        item_num += 1

        # Parse review flag from notes
        notes = r['notes'] or ''
        flag = ''
        clean_notes = notes
        if '[REVIEW-SUMMARIZE]' in notes:
            flag = 'Summarize?'
            clean_notes = notes.replace('[REVIEW-SUMMARIZE]', '').strip()
        elif '[REVIEW-NO-QTY]' in notes:
            flag = 'No Qty'
            clean_notes = notes.replace('[REVIEW-NO-QTY]', '').strip()

        # Combine model_number and specs
        specs = r['specifications'] or ''

        values = [
            item_num,
            r['category'],
            r['description'],
            r['brand'] or '',
            specs,
            r['expected_qty'],
            r['unit'] or 'EA',
            'Yes' if r['is_bookable'] else '',
            flag,
            clean_notes,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin_border
            cell.alignment = wrap if col in (5, 10) else (center if col in (1, 6, 7, 8) else Alignment(vertical="top"))

            # Highlight bookable rows
            if r['is_bookable'] and col not in (1,):
                cell.fill = bookable_fill
            # Highlight flagged rows
            if flag and col == 9:
                cell.fill = flag_fill
                cell.font = Font(bold=True, color="B45309")

    # ── Summary sheet ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12

    sum_headers = ["Category", "Models", "Total Units"]
    for col, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    with get_db() as conn:
        summary = conn.execute("""
            SELECT c.name, COUNT(*) as models,
                   COALESCE(SUM(em.expected_qty), 0) as units
            FROM equipment_models em
            JOIN categories c ON em.category_id = c.id
            GROUP BY c.name ORDER BY units DESC
        """).fetchall()

    for i, s in enumerate(summary, 2):
        ws2.cell(row=i, column=1, value=s['name'])
        ws2.cell(row=i, column=2, value=s['models']).alignment = center
        ws2.cell(row=i, column=3, value=s['units']).alignment = center

    total_row = len(summary) + 2
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=2,
             value=sum(s['models'] for s in summary)).font = Font(bold=True)
    ws2.cell(row=total_row, column=3,
             value=sum(s['units'] for s in summary)).font = Font(bold=True)

    ws2.freeze_panes = "A2"

    # ── Save ─────────────────────────────────────────────────────────
    wb.save(OUT)
    print(f"Exported {item_num} items to {OUT}")
    print(f"  Sheet 1: Equipment Inventory (with category groupings)")
    print(f"  Sheet 2: Summary by category")


if __name__ == "__main__":
    main()
