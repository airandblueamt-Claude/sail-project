"""Daily Inspection Checklist — lab cleanliness / operational round.

One inspection per calendar day. Items live in the inspection_items catalog,
keyed by area. The Today view is the team's daily workflow: open, click
Active/Inactive/None per item, sign off, done. The dashboard surfaces
"what's broken" so the team can watch lab health over time.
"""
import calendar
import io
import os
import uuid
from datetime import date, datetime, timedelta, timezone

from flask import (Blueprint, current_app, g, jsonify, redirect,
                   render_template, request, url_for, flash, abort, Response)

from database import get_db, log_audit
from email_service import send_email_with_attachment, is_email_configured

# Image types accepted for an issue photo (mirrors the inventory uploader).
_PHOTO_EXT = {'png', 'jpg', 'jpeg', 'webp', 'gif', 'heic'}


def _save_issue_photo(file):
    """Save an uploaded issue photo; return 'uploads/<name>' or None."""
    if not file or not file.filename or '.' not in file.filename:
        return None
    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in _PHOTO_EXT:
        return None
    filename = f"insp_{uuid.uuid4().hex[:12]}.{ext}"
    file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
    return f"uploads/{filename}"

try:
    from zoneinfo import ZoneInfo
    _LOCAL_TZ = ZoneInfo('Asia/Riyadh')
except Exception:  # pragma: no cover
    _LOCAL_TZ = None


def _now_local_str():
    """Wall-clock 'YYYY-MM-DD HH:MM' for the team's timezone."""
    now = datetime.now(timezone.utc)
    if _LOCAL_TZ is not None:
        now = now.astimezone(_LOCAL_TZ)
    return now.strftime('%Y-%m-%d %H:%M')

inspections_bp = Blueprint('inspections', __name__)

STATUSES = ('active', 'inactive', 'none')
STAFF_ROLES = ('admin', 'manager', 'technician')
ADMIN_ROLES = ('admin',)


# ── small helpers ──────────────────────────────────────────────────────────

def _is_staff():
    return g.user and g.user['role'] in STAFF_ROLES


def _is_admin():
    return g.user and g.user['role'] in ADMIN_ROLES


# Section display order on the inspection page. Areas whose `section` is empty
# or not listed here fall through to "Other" at the end (see _group_into_sections).
SECTION_ORDER = [
    "Facilities & Infrastructure",
    "Incubators",
    "Labs",
    "Workshops & Studios",
    "Operations & Offices",
]
OTHER_SECTION = "Other"


def _load_catalog(conn):
    """Return [{area: {...}, items: [{...}]}, ...] ordered for rendering."""
    areas = conn.execute(
        "SELECT id, name, display_order, section FROM inspection_areas "
        "WHERE is_active = 1 ORDER BY display_order, name"
    ).fetchall()
    items = conn.execute(
        "SELECT id, area_id, name, display_order, is_applicable "
        "FROM inspection_items "
        "WHERE is_active = 1 ORDER BY area_id, display_order, name"
    ).fetchall()
    by_area = {}
    for it in items:
        by_area.setdefault(it['area_id'], []).append(dict(it))
    return [{'area': dict(a), 'items': by_area.get(a['id'], [])} for a in areas]


def _group_into_sections(catalog):
    """Group a flat catalog into ordered sections for the page.

    Returns [{name, groups, item_count}, ...]. Section order follows
    SECTION_ORDER; any area with an unknown/empty section lands in "Other",
    appended last. Within a section, areas keep their catalog (display_order)
    order. Empty sections are dropped.
    """
    buckets = {}
    for grp in catalog:
        label = (grp['area'].get('section') or '').strip() or OTHER_SECTION
        buckets.setdefault(label, []).append(grp)
    ordered_labels = [s for s in SECTION_ORDER if s in buckets]
    ordered_labels += [s for s in buckets if s not in SECTION_ORDER]  # incl. Other
    sections = []
    for label in ordered_labels:
        groups = buckets[label]
        sections.append({
            'name': label,
            'groups': groups,
            'item_count': sum(1 for g in groups for it in g['items']
                              if it.get('is_applicable', 1)),
        })
    return sections


# ── Sheets: two owned halves of the daily round (Infra / Rooms) ──────────────
# The map from section -> sheet lives here (the DB only stores order + assignee).
SHEET_OF_SECTION = {
    "Facilities & Infrastructure": "Infra",
    "Incubators":          "Rooms",
    "Labs":                "Rooms",
    "Workshops & Studios": "Rooms",
    "Operations & Offices": "Rooms",
}
DEFAULT_SHEET = "Rooms"   # any unmapped section falls here


def _sheet_for_section(section_name):
    return SHEET_OF_SECTION.get(section_name, DEFAULT_SHEET)


def _sheet_meta(conn):
    """{name: {display_order, assignee_id, assignee_name}} for both sheets."""
    return {r['name']: dict(r) for r in conn.execute(
        """SELECT s.name, s.display_order, s.assignee_id, e.name AS assignee_name
           FROM inspection_sheets s
           LEFT JOIN employees e ON e.id = s.assignee_id""").fetchall()}


def _build_sheets(conn, catalog, inspection_id, results):
    """Group sections into the two sheets, with per-sheet progress + sign-off.

    Returns [{name, assignee_id, assignee_name, sections, total, recorded,
              inactive, pct, area_ids, signoff, order}, ...].
    """
    sections = _group_into_sections(catalog)
    meta = _sheet_meta(conn)
    signoffs = {}
    if inspection_id:
        for r in conn.execute(
            """SELECT so.sheet, so.signed_at, so.signed_by, e.name AS signer_name
               FROM inspection_sheet_signoffs so
               LEFT JOIN employees e ON e.id = so.signed_by
               WHERE so.inspection_id = ?""", (inspection_id,)):
            signoffs[r['sheet']] = dict(r)

    buckets = {}
    for sec in sections:
        buckets.setdefault(_sheet_for_section(sec['name']), []).append(sec)

    sheets = []
    for name, secs in buckets.items():
        total = recorded = inactive = 0
        area_ids = []
        for sec in secs:
            for g in sec['groups']:
                area_ids.append(g['area']['id'])
                for it in g['items']:
                    if not it.get('is_applicable', 1):
                        continue
                    total += 1
                    r = results.get(it['id'])
                    if r:
                        recorded += 1
                        if r.get('status') == 'inactive':
                            inactive += 1
        m = meta.get(name, {})
        sheets.append({
            'name': name,
            'sections': secs,
            'assignee_id': m.get('assignee_id'),
            'assignee_name': m.get('assignee_name'),
            'total': total, 'recorded': recorded, 'inactive': inactive,
            'pct': round(100 * recorded / total) if total else 0,
            'area_ids': area_ids,
            'signoff': signoffs.get(name),
            'order': m.get('display_order', 999),
        })
    sheets.sort(key=lambda s: (s['order'], s['name']))
    return sheets


def _load_results(conn, inspection_id):
    """Return {item_id: result_dict} for an inspection."""
    if not inspection_id:
        return {}
    rows = conn.execute(
        "SELECT id, item_id, status, notes, photo_path, updated_at, updated_by "
        "FROM inspection_results WHERE inspection_id = ?",
        (inspection_id,)
    ).fetchall()
    return {r['item_id']: dict(r) for r in rows}


def _inspection_with_signatures(conn, inspection_id):
    return conn.execute(
        """SELECT i.*,
                  c.name AS creator_name,
                  ie.name AS engineer_name,
                  ams.name AS amt_supervisor_name,
                  ss.name AS sail_supervisor_name,
                  hd.name AS head_name
           FROM inspections i
           LEFT JOIN employees c   ON i.created_by             = c.id
           LEFT JOIN employees ie  ON i.inspection_engineer_id = ie.id
           LEFT JOIN employees ams ON i.amt_supervisor_id      = ams.id
           LEFT JOIN employees ss  ON i.sail_supervisor_id     = ss.id
           LEFT JOIN employees hd  ON i.head_id                = hd.id
           WHERE i.id = ?""",
        (inspection_id,)
    ).fetchone()


def _completion(conn, inspection_id, catalog):
    """Return (total_items, recorded, inactive_count) for an inspection.

    N/A items (is_applicable = 0) are excluded from both the denominator and
    any stray recorded rows, so a day with every applicable item set reads 100%.
    """
    total = sum(1 for g in catalog for it in g['items']
                if it.get('is_applicable', 1))
    if not inspection_id or total == 0:
        return total, 0, 0
    row = conn.execute(
        """SELECT
              COUNT(*) AS recorded,
              SUM(CASE WHEN r.status = 'inactive' THEN 1 ELSE 0 END) AS inactive
           FROM inspection_results r
           JOIN inspection_items it ON it.id = r.item_id
           WHERE r.inspection_id = ?
             AND it.is_active = 1 AND it.is_applicable = 1""",
        (inspection_id,)
    ).fetchone()
    return total, int(row['recorded'] or 0), int(row['inactive'] or 0)


# ── dashboard ──────────────────────────────────────────────────────────────

@inspections_bp.route('/')
def dashboard():
    today_str = date.today().isoformat()
    thirty_days_ago = (date.today() - timedelta(days=29)).isoformat()
    seven_days_ago  = (date.today() - timedelta(days=6)).isoformat()

    with get_db() as conn:
        catalog = _load_catalog(conn)
        total_items = sum(1 for g in catalog for it in g['items']
                          if it.get('is_applicable', 1))

        today_row = conn.execute(
            """SELECT i.*,
                      hd.name AS head_name
               FROM inspections i
               LEFT JOIN employees hd ON i.head_id = hd.id
               WHERE i.inspection_date = ?""",
            (today_str,)
        ).fetchone()
        _, recorded, inactive_today = _completion(
            conn, today_row['id'] if today_row else None, catalog)
        completion_pct = round(100 * recorded / total_items) if total_items else 0

        # 30-day trend: one row per inspection, with inactive counts.
        trend_rows = conn.execute(
            """SELECT i.inspection_date,
                      COUNT(r.id) AS recorded,
                      SUM(CASE WHEN r.status='inactive' THEN 1 ELSE 0 END) AS inactive
               FROM inspections i
               LEFT JOIN inspection_results r ON r.inspection_id = i.id
               WHERE i.inspection_date >= ?
               GROUP BY i.id
               ORDER BY i.inspection_date""",
            (thirty_days_ago,)
        ).fetchall()
        trend_by_date = {r['inspection_date']: dict(r) for r in trend_rows}
        max_inactive = max(
            (r['inactive'] or 0 for r in trend_rows), default=0)
        trend = []
        for offset in range(30):
            d = (date.today() - timedelta(days=29 - offset)).isoformat()
            r = trend_by_date.get(d)
            trend.append({
                'date': d,
                'inactive': int(r['inactive']) if r and r['inactive'] else 0,
                'recorded': int(r['recorded']) if r else 0,
                'height_pct': round(100 * (r['inactive'] or 0) / max_inactive)
                              if (r and max_inactive) else 0,
                'has_row': bool(r),
            })

        # 7-day submitted streak (consecutive days from today backwards).
        submitted_dates = {r[0] for r in conn.execute(
            "SELECT inspection_date FROM inspections "
            "WHERE submitted_at IS NOT NULL AND inspection_date >= ?",
            (seven_days_ago,)
        ).fetchall()}
        streak = 0
        for offset in range(0, 7):
            d = (date.today() - timedelta(days=offset)).isoformat()
            if d in submitted_dates:
                streak += 1
            else:
                break

        # Top recurring problems (Inactive in last 30 days, top 10).
        recurring = conn.execute(
            """SELECT it.id   AS item_id, it.name AS item_name,
                      ar.name AS area_name,
                      COUNT(*) AS hits
               FROM inspection_results r
               JOIN inspection_items it ON r.item_id = it.id
               JOIN inspection_areas ar ON it.area_id = ar.id
               JOIN inspections i        ON r.inspection_id = i.id
               WHERE r.status = 'inactive'
                 AND i.inspection_date >= ?
               GROUP BY it.id
               ORDER BY hits DESC, item_name
               LIMIT 10""",
            (thirty_days_ago,)
        ).fetchall()

        # Last 7 inspections strip.
        last7 = conn.execute(
            """SELECT i.id, i.inspection_date, i.submitted_at,
                      e.name AS engineer_name,
                      COUNT(r.id) AS recorded,
                      SUM(CASE WHEN r.status='inactive' THEN 1 ELSE 0 END) AS inactive
               FROM inspections i
               LEFT JOIN employees e ON i.inspection_engineer_id = e.id
               LEFT JOIN inspection_results r ON r.inspection_id = i.id
               GROUP BY i.id
               ORDER BY i.inspection_date DESC
               LIMIT 7"""
        ).fetchall()

        # Per-area summary for the "Today's status" panel.
        area_status = []
        results_today = _load_results(
            conn, today_row['id'] if today_row else None)
        for grp in catalog:
            items = [it for it in grp['items'] if it.get('is_applicable', 1)]
            recorded_n = sum(1 for it in items if it['id'] in results_today)
            inactive_n = sum(
                1 for it in items
                if results_today.get(it['id'], {}).get('status') == 'inactive')
            area_status.append({
                'area': grp['area'],
                'item_count': len(items),
                'recorded': recorded_n,
                'inactive': inactive_n,
            })

        # Per-sheet (Infra / Rooms) rollup with assignee + sign-off, for tracking.
        sheets_today = _build_sheets(
            conn, catalog, today_row['id'] if today_row else None, results_today)

    return render_template(
        'inspections/dashboard.html',
        today=today_row,
        today_str=today_str,
        total_items=total_items,
        recorded_today=recorded,
        inactive_today=inactive_today,
        completion_pct=completion_pct,
        pending_areas=[a for a in area_status if a['recorded'] == 0],
        area_status=area_status,
        sheets_today=sheets_today,
        trend=trend,
        max_inactive=max_inactive,
        recurring=recurring,
        last7=last7,
        streak=streak,
        is_staff=_is_staff(),
    )


# ── today / detail views ──────────────────────────────────────────────────

@inspections_bp.route('/today')
def today():
    today_str = date.today().isoformat()
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM inspections WHERE inspection_date = ?",
            (today_str,)
        ).fetchone()
        if not row and _is_staff():
            cur = conn.execute(
                "INSERT INTO inspections (inspection_date, created_by) "
                "VALUES (?, ?)",
                (today_str, g.user['id']))
            log_audit(conn, 'inspections', cur.lastrowid, 'create',
                      changed_by=g.user['id'])
            row = conn.execute(
                "SELECT * FROM inspections WHERE id = ?",
                (cur.lastrowid,)).fetchone()

        inspection = _inspection_with_signatures(conn, row['id']) if row else None
        catalog = _load_catalog(conn)
        results = _load_results(conn, row['id'] if row else None)
        total, recorded, inactive = _completion(
            conn, row['id'] if row else None, catalog)
        sheets = _build_sheets(conn, catalog, row['id'] if row else None, results)

        signatories = conn.execute(
            "SELECT id, name, role FROM employees "
            "WHERE is_active = 1 ORDER BY name"
        ).fetchall()

    editable = bool(inspection) and _is_staff() and (
        not inspection['submitted_at'] or _is_admin())

    return render_template(
        'inspections/detail.html',
        inspection=inspection,
        catalog=catalog,
        sheets=sheets,
        results=results,
        signatories=signatories,
        total_items=total,
        recorded=recorded,
        inactive=inactive,
        completion_pct=round(100 * recorded / total) if total else 0,
        editable=editable,
        is_admin=_is_admin(),
        email_ready=is_email_configured(),
        view_mode='today',
        view_date=date.today(),
    )


@inspections_bp.route('/<inspection_date>')
def detail(inspection_date):
    # Validate the date so we don't accept random strings as ids.
    try:
        view_date = datetime.strptime(inspection_date, '%Y-%m-%d').date()
    except ValueError:
        abort(404)

    if view_date == date.today():
        return redirect(url_for('inspections.today'))

    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM inspections WHERE inspection_date = ?",
            (inspection_date,)
        ).fetchone()
        if not row:
            flash(f'No inspection recorded for {inspection_date}.', 'info')
            return redirect(url_for('inspections.history'))

        inspection = _inspection_with_signatures(conn, row['id'])
        catalog = _load_catalog(conn)
        results = _load_results(conn, row['id'])
        total, recorded, inactive = _completion(conn, row['id'], catalog)
        sheets = _build_sheets(conn, catalog, row['id'], results)

        signatories = conn.execute(
            "SELECT id, name, role FROM employees "
            "WHERE is_active = 1 ORDER BY name"
        ).fetchall()

    editable = _is_admin() or (
        _is_staff() and not inspection['submitted_at'])

    return render_template(
        'inspections/detail.html',
        inspection=inspection,
        catalog=catalog,
        sheets=sheets,
        results=results,
        signatories=signatories,
        total_items=total,
        recorded=recorded,
        inactive=inactive,
        completion_pct=round(100 * recorded / total) if total else 0,
        editable=editable,
        is_admin=_is_admin(),
        email_ready=is_email_configured(),
        view_mode='detail',
        view_date=view_date,
    )


# ── item updates (AJAX) ────────────────────────────────────────────────────

@inspections_bp.route('/<int:inspection_id>/items/<int:item_id>',
                      methods=['POST'])
def set_item(inspection_id, item_id):
    if not _is_staff():
        return jsonify(error='forbidden',
                       message='Only staff may record inspection items.'), 403

    new_status = (request.json or request.form or {}).get('status', '').strip()
    if new_status not in STATUSES:
        return jsonify(error='bad_request',
                       message='status must be active, inactive, or none.'), 400

    with get_db() as conn:
        insp = conn.execute(
            "SELECT id, submitted_at FROM inspections WHERE id = ?",
            (inspection_id,)
        ).fetchone()
        if not insp:
            return jsonify(error='not_found'), 404
        if insp['submitted_at'] and not _is_admin():
            return jsonify(
                error='locked',
                message='Inspection has been submitted; '
                        'only admins can edit.'), 409

        # Item must exist, be active, and be applicable (N/A items aren't recorded).
        item = conn.execute(
            "SELECT id FROM inspection_items "
            "WHERE id = ? AND is_active = 1 AND is_applicable = 1", (item_id,)
        ).fetchone()
        if not item:
            return jsonify(error='not_found',
                           message='Item not found, inactive, or N/A.'), 404

        existing = conn.execute(
            "SELECT id, status FROM inspection_results "
            "WHERE inspection_id = ? AND item_id = ?",
            (inspection_id, item_id)
        ).fetchone()

        if existing:
            conn.execute(
                "UPDATE inspection_results "
                "SET status = ?, updated_by = ?, updated_at = datetime('now') "
                "WHERE id = ?",
                (new_status, g.user['id'], existing['id']))
            log_audit(conn, 'inspection_results', existing['id'],
                      'status_change', 'status',
                      existing['status'], new_status,
                      changed_by=g.user['id'])
        else:
            cur = conn.execute(
                "INSERT INTO inspection_results "
                "(inspection_id, item_id, status, updated_by) "
                "VALUES (?, ?, ?, ?)",
                (inspection_id, item_id, new_status, g.user['id']))
            log_audit(conn, 'inspection_results', cur.lastrowid, 'create',
                      'status', None, new_status,
                      changed_by=g.user['id'])

        conn.execute(
            "UPDATE inspections SET updated_at = datetime('now') WHERE id = ?",
            (inspection_id,))

        stats = _inspection_stats(conn, inspection_id)

    return jsonify(
        ok=True,
        item_id=item_id,
        status=new_status,
        recorded=int(stats['recorded'] or 0),
        total=int(stats['total'] or 0),
        inactive=int(stats['inactive'] or 0),
        completion_pct=round(100 * (stats['recorded'] or 0) /
                             (stats['total'] or 1)),
        updated_at=datetime.utcnow().isoformat() + 'Z',
        updated_at_local=_now_local_str(),
    )


@inspections_bp.route('/<int:inspection_id>/items/<int:item_id>/note',
                      methods=['POST'])
def set_item_note(inspection_id, item_id):
    if not _is_staff():
        return jsonify(error='forbidden'), 403
    note = (request.json or request.form or {}).get('notes', '')
    note = (note or '').strip()[:500]

    with get_db() as conn:
        insp = conn.execute(
            "SELECT id, submitted_at FROM inspections WHERE id = ?",
            (inspection_id,)).fetchone()
        if not insp:
            return jsonify(error='not_found'), 404
        if insp['submitted_at'] and not _is_admin():
            return jsonify(error='locked'), 409

        existing = conn.execute(
            "SELECT id FROM inspection_results "
            "WHERE inspection_id = ? AND item_id = ?",
            (inspection_id, item_id)).fetchone()
        if existing:
            conn.execute(
                "UPDATE inspection_results SET notes = ?, "
                "updated_at = datetime('now'), updated_by = ? WHERE id = ?",
                (note or None, g.user['id'], existing['id']))
        else:
            # No status yet → default to 'none' so a bare note is still valid.
            conn.execute(
                "INSERT INTO inspection_results "
                "(inspection_id, item_id, status, notes, updated_by) "
                "VALUES (?, ?, 'none', ?, ?)",
                (inspection_id, item_id, note or None, g.user['id']))

    return jsonify(ok=True, notes=note)


@inspections_bp.route('/<int:inspection_id>/items/<int:item_id>/photo',
                      methods=['POST'])
def set_item_photo(inspection_id, item_id):
    """Attach (or, with remove=1, clear) an evidence photo on an item."""
    if not _is_staff():
        return jsonify(error='forbidden'), 403

    with get_db() as conn:
        insp = conn.execute(
            "SELECT id, submitted_at FROM inspections WHERE id = ?",
            (inspection_id,)).fetchone()
        if not insp:
            return jsonify(error='not_found'), 404
        if insp['submitted_at'] and not _is_admin():
            return jsonify(error='locked'), 409

        existing = conn.execute(
            "SELECT id, photo_path FROM inspection_results "
            "WHERE inspection_id = ? AND item_id = ?",
            (inspection_id, item_id)).fetchone()

        # Remove path.
        remove_flag = request.form.get('remove') or (
            request.is_json and (request.json or {}).get('remove'))
        if remove_flag:
            if existing and existing['photo_path']:
                _delete_photo_file(existing['photo_path'])
                conn.execute(
                    "UPDATE inspection_results SET photo_path = NULL, "
                    "updated_at = datetime('now'), updated_by = ? WHERE id = ?",
                    (g.user['id'], existing['id']))
            return jsonify(ok=True, photo_url=None)

        rel = _save_issue_photo(request.files.get('photo'))
        if not rel:
            return jsonify(error='bad_request',
                           message='No valid image (png/jpg/webp/gif/heic).'), 400

        if existing:
            if existing['photo_path']:
                _delete_photo_file(existing['photo_path'])  # replace old file
            conn.execute(
                "UPDATE inspection_results SET photo_path = ?, "
                "updated_at = datetime('now'), updated_by = ? WHERE id = ?",
                (rel, g.user['id'], existing['id']))
        else:
            # No status yet → default to 'none' so the photo has a row to live on.
            conn.execute(
                "INSERT INTO inspection_results "
                "(inspection_id, item_id, status, photo_path, updated_by) "
                "VALUES (?, ?, 'none', ?, ?)",
                (inspection_id, item_id, rel, g.user['id']))

    return jsonify(ok=True, photo_url=url_for('static', filename=rel))


def _delete_photo_file(rel_path):
    """Best-effort removal of a stored photo file (rel = 'uploads/<name>')."""
    try:
        if rel_path and rel_path.startswith('uploads/'):
            name = rel_path.split('/', 1)[1]
            fp = os.path.join(current_app.config['UPLOAD_FOLDER'], name)
            if os.path.isfile(fp):
                os.remove(fp)
    except Exception:
        pass


def _inspection_stats(conn, inspection_id):
    """Shared (total, recorded, inactive) snapshot for AJAX responses.

    Counts only applicable items, and only results that belong to applicable
    items, so completion can reach 100% with N/A items present.
    """
    return conn.execute(
        """SELECT
              (SELECT COUNT(*) FROM inspection_items
                 WHERE is_active = 1 AND is_applicable = 1) AS total,
              (SELECT COUNT(*) FROM inspection_results r
                 JOIN inspection_items it ON it.id = r.item_id
                 WHERE r.inspection_id = ?
                   AND it.is_active = 1 AND it.is_applicable = 1) AS recorded,
              (SELECT COUNT(*) FROM inspection_results r
                 JOIN inspection_items it ON it.id = r.item_id
                 WHERE r.inspection_id = ? AND r.status = 'inactive'
                   AND it.is_active = 1 AND it.is_applicable = 1) AS inactive""",
        (inspection_id, inspection_id)).fetchone()


@inspections_bp.route('/<int:inspection_id>/bulk', methods=['POST'])
def bulk_set(inspection_id):
    """Set many items at once — one tap for a whole area, section, or the page.

    Body: {status, area_ids?: [int], only_unrecorded?: bool (default true)}.
    only_unrecorded leaves already-recorded items untouched, so an inspector can
    flag the exceptions first and then "fill the rest" without clobbering them.
    """
    if not _is_staff():
        return jsonify(error='forbidden',
                       message='Only staff may record inspection items.'), 403
    body = request.json or request.form or {}
    status = (body.get('status') or '').strip()
    if status not in STATUSES:
        return jsonify(error='bad_request',
                       message='status must be active, inactive, or none.'), 400
    only_unrecorded = bool(body.get('only_unrecorded', True))
    area_ids = body.get('area_ids')
    if area_ids is not None:
        try:
            area_ids = [int(a) for a in area_ids]
        except (TypeError, ValueError):
            return jsonify(error='bad_request',
                           message='area_ids must be integers.'), 400

    with get_db() as conn:
        insp = conn.execute(
            "SELECT id, submitted_at FROM inspections WHERE id = ?",
            (inspection_id,)).fetchone()
        if not insp:
            return jsonify(error='not_found'), 404
        if insp['submitted_at'] and not _is_admin():
            return jsonify(error='locked',
                           message='Inspection has been submitted; '
                                   'only admins can edit.'), 409

        q = "SELECT id FROM inspection_items WHERE is_active = 1 AND is_applicable = 1"
        params = []
        if area_ids is not None:
            if not area_ids:           # explicit empty scope → nothing to do
                stats = _inspection_stats(conn, inspection_id)
                return jsonify(ok=True, set=0, item_ids=[], status=status,
                               recorded=int(stats['recorded'] or 0),
                               total=int(stats['total'] or 0),
                               inactive=int(stats['inactive'] or 0),
                               completion_pct=round(100 * (stats['recorded'] or 0)
                                                    / (stats['total'] or 1)))
            q += " AND area_id IN (%s)" % ",".join("?" * len(area_ids))
            params += area_ids
        if only_unrecorded:
            q += (" AND id NOT IN (SELECT item_id FROM inspection_results "
                  "WHERE inspection_id = ?)")
            params.append(inspection_id)
        targets = [r['id'] for r in conn.execute(q, params).fetchall()]

        for iid in targets:
            conn.execute(
                "INSERT INTO inspection_results "
                "(inspection_id, item_id, status, updated_by) "
                "VALUES (?, ?, ?, ?) "
                "ON CONFLICT(inspection_id, item_id) DO UPDATE SET "
                "  status = excluded.status, updated_by = excluded.updated_by, "
                "  updated_at = datetime('now')",
                (inspection_id, iid, status, g.user['id']))
        if targets:
            log_audit(conn, 'inspections', inspection_id, 'status_change',
                      'bulk_active', None, f'{status} ({len(targets)} items)',
                      changed_by=g.user['id'])
            conn.execute(
                "UPDATE inspections SET updated_at = datetime('now') WHERE id = ?",
                (inspection_id,))

        stats = _inspection_stats(conn, inspection_id)

    return jsonify(
        ok=True,
        set=len(targets),
        item_ids=targets,
        status=status,
        recorded=int(stats['recorded'] or 0),
        total=int(stats['total'] or 0),
        inactive=int(stats['inactive'] or 0),
        completion_pct=round(100 * (stats['recorded'] or 0) /
                             (stats['total'] or 1)),
        updated_at_local=_now_local_str(),
    )


@inspections_bp.route('/<int:inspection_id>/prefill', methods=['POST'])
def prefill(inspection_id):
    """Copy the most recent earlier inspection's results into this one.

    Only fills items not yet recorded today (won't clobber edits), and only
    items that still exist and are active. Plain form POST → redirect.
    """
    if not _is_staff():
        flash('Only staff may record inspections.', 'error')
        return redirect(url_for('inspections.today'))
    with get_db() as conn:
        insp = conn.execute(
            "SELECT id, inspection_date, submitted_at "
            "FROM inspections WHERE id = ?", (inspection_id,)).fetchone()
        if not insp:
            abort(404)
        if insp['submitted_at'] and not _is_admin():
            flash('Inspection has been submitted; only admins can edit.', 'error')
            return redirect(url_for('inspections.today'))

        src = conn.execute(
            """SELECT i.id, i.inspection_date
               FROM inspections i
               WHERE i.inspection_date < ?
                 AND EXISTS (SELECT 1 FROM inspection_results r
                             WHERE r.inspection_id = i.id)
               ORDER BY i.inspection_date DESC
               LIMIT 1""",
            (insp['inspection_date'],)).fetchone()
        if not src:
            flash('No earlier inspection to copy from.', 'info')
            return redirect(url_for('inspections.today'))

        rows = conn.execute(
            """SELECT r.item_id, r.status, r.notes
               FROM inspection_results r
               JOIN inspection_items it ON it.id = r.item_id AND it.is_active = 1
               WHERE r.inspection_id = ?
                 AND r.item_id NOT IN (
                     SELECT item_id FROM inspection_results
                     WHERE inspection_id = ?)""",
            (src['id'], inspection_id)).fetchall()
        n = 0
        for r in rows:
            conn.execute(
                "INSERT INTO inspection_results "
                "(inspection_id, item_id, status, notes, updated_by) "
                "VALUES (?, ?, ?, ?, ?) "
                "ON CONFLICT(inspection_id, item_id) DO NOTHING",
                (inspection_id, r['item_id'], r['status'], r['notes'],
                 g.user['id']))
            n += 1
        if n:
            log_audit(conn, 'inspections', inspection_id, 'update', 'prefill',
                      None, f"from {src['inspection_date']} ({n} items)",
                      changed_by=g.user['id'])
            conn.execute(
                "UPDATE inspections SET updated_at = datetime('now') WHERE id = ?",
                (inspection_id,))
    flash(f"Prefilled {n} item(s) from {src['inspection_date']}. "
          f"Review the round and edit any exceptions.", 'success')
    return redirect(url_for('inspections.today'))


# ── submit / reopen ────────────────────────────────────────────────────────

@inspections_bp.route('/<int:inspection_id>/submit', methods=['POST'])
def submit(inspection_id):
    if not _is_staff():
        flash('Only staff may submit inspections.', 'error')
        return redirect(url_for('inspections.today'))

    def _int_or_none(v):
        try:
            return int(v) if v not in (None, '') else None
        except (TypeError, ValueError):
            return None

    head = _int_or_none(request.form.get('head_id'))
    notes    = (request.form.get('notes') or '').strip()[:2000]

    with get_db() as conn:
        insp = conn.execute(
            "SELECT id, submitted_at, inspection_date FROM inspections "
            "WHERE id = ?", (inspection_id,)).fetchone()
        if not insp:
            flash('Inspection not found.', 'error')
            return redirect(url_for('inspections.dashboard'))
        if insp['submitted_at'] and not _is_admin():
            flash('Inspection is already submitted; only admins can change it.',
                  'error')
            return redirect(url_for('inspections.detail',
                                    inspection_date=insp['inspection_date']))
        if not head:
            flash('Select the Head signing off before submitting.', 'error')
            return redirect(url_for('inspections.detail',
                                    inspection_date=insp['inspection_date'])
                            if insp['inspection_date'] != date.today().isoformat()
                            else url_for('inspections.today'))

        conn.execute(
            """UPDATE inspections SET
                 head_id      = ?,
                 notes        = ?,
                 submitted_at = COALESCE(submitted_at, datetime('now')),
                 updated_at   = datetime('now')
               WHERE id = ?""",
            (head, notes or None, inspection_id))
        log_audit(conn, 'inspections', inspection_id,
                  'status_change', 'submitted_at',
                  insp['submitted_at'], 'now', changed_by=g.user['id'])

    flash('Inspection signed off & submitted.', 'success')
    today_str = date.today().isoformat()
    if insp['inspection_date'] == today_str:
        return redirect(url_for('inspections.today'))
    return redirect(url_for('inspections.detail',
                            inspection_date=insp['inspection_date']))


@inspections_bp.route('/<int:inspection_id>/reopen', methods=['POST'])
def reopen(inspection_id):
    if not _is_admin():
        flash('Only admins can reopen a submitted inspection.', 'error')
        return redirect(url_for('inspections.dashboard'))
    with get_db() as conn:
        insp = conn.execute(
            "SELECT id, submitted_at, inspection_date FROM inspections "
            "WHERE id = ?", (inspection_id,)).fetchone()
        if not insp:
            flash('Inspection not found.', 'error')
            return redirect(url_for('inspections.dashboard'))
        conn.execute(
            "UPDATE inspections SET submitted_at = NULL, "
            "updated_at = datetime('now') WHERE id = ?", (inspection_id,))
        log_audit(conn, 'inspections', inspection_id,
                  'status_change', 'submitted_at',
                  insp['submitted_at'], None, changed_by=g.user['id'])
    flash('Inspection reopened for editing.', 'success')
    today_str = date.today().isoformat()
    if insp['inspection_date'] == today_str:
        return redirect(url_for('inspections.today'))
    return redirect(url_for('inspections.detail',
                            inspection_date=insp['inspection_date']))


# ── per-sheet sign-off ──────────────────────────────────────────────────────

@inspections_bp.route('/<int:inspection_id>/sheet-signoff', methods=['POST'])
def sheet_signoff(inspection_id):
    """Toggle one sheet's (Infra/Rooms) sign-off for an inspection.

    Soft model: signing doesn't lock editing — it records that the responsible
    person confirms their half is done. The signer or an admin can remove it.
    """
    if not _is_staff():
        flash('Only staff can sign off a sheet.', 'error')
        return redirect(url_for('inspections.today'))
    sheet = (request.form.get('sheet') or '').strip()
    with get_db() as conn:
        insp = conn.execute(
            "SELECT id, inspection_date, submitted_at FROM inspections "
            "WHERE id = ?", (inspection_id,)).fetchone()
        if not insp:
            abort(404)
        known = {r[0] for r in conn.execute(
            "SELECT name FROM inspection_sheets").fetchall()}
        if sheet not in known:
            flash('Unknown sheet.', 'error')
        elif insp['submitted_at'] and not _is_admin():
            flash('Inspection is submitted; only admins can change sign-off.',
                  'error')
        else:
            existing = conn.execute(
                "SELECT id, signed_by FROM inspection_sheet_signoffs "
                "WHERE inspection_id = ? AND sheet = ?",
                (inspection_id, sheet)).fetchone()
            if existing:
                if _is_admin() or existing['signed_by'] == g.user['id']:
                    conn.execute(
                        "DELETE FROM inspection_sheet_signoffs WHERE id = ?",
                        (existing['id'],))
                    log_audit(conn, 'inspection_sheet_signoffs', existing['id'],
                              'delete', 'sheet', sheet, None,
                              changed_by=g.user['id'])
                    flash(f'{sheet} marked not complete.', 'info')
                else:
                    flash('Only the person who marked it, or an admin, can undo.',
                          'error')
            else:
                cur = conn.execute(
                    "INSERT INTO inspection_sheet_signoffs "
                    "(inspection_id, sheet, signed_by) VALUES (?, ?, ?)",
                    (inspection_id, sheet, g.user['id']))
                log_audit(conn, 'inspection_sheet_signoffs', cur.lastrowid,
                          'create', 'sheet', None, sheet,
                          changed_by=g.user['id'])
                flash(f'{sheet} marked complete — thank you.', 'success')
    if insp['inspection_date'] == date.today().isoformat():
        return redirect(url_for('inspections.today'))
    return redirect(url_for('inspections.detail',
                            inspection_date=insp['inspection_date']))


# ── history ────────────────────────────────────────────────────────────────

@inspections_bp.route('/history')
def history():
    from_date = (request.args.get('from') or '').strip()
    to_date   = (request.args.get('to') or '').strip()
    has_inactive = request.args.get('has_inactive') in ('1', 'true', 'yes', 'on')

    where, params = [], []
    if from_date:
        where.append('i.inspection_date >= ?'); params.append(from_date)
    if to_date:
        where.append('i.inspection_date <= ?'); params.append(to_date)
    where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''

    with get_db() as conn:
        rows = conn.execute(
            f"""SELECT i.id, i.inspection_date, i.submitted_at,
                       i.created_at,
                       cr.name AS creator_name,
                       hd.name AS head_name,
                       (SELECT COUNT(*) FROM inspection_results r
                          JOIN inspection_items it ON it.id = r.item_id
                          WHERE r.inspection_id = i.id
                            AND it.is_active = 1 AND it.is_applicable = 1) AS recorded,
                       (SELECT COUNT(*) FROM inspection_results r
                          JOIN inspection_items it ON it.id = r.item_id
                          WHERE r.inspection_id = i.id AND r.status = 'inactive'
                            AND it.is_active = 1 AND it.is_applicable = 1) AS inactive,
                       (SELECT COUNT(*) FROM inspection_items
                          WHERE is_active = 1 AND is_applicable = 1) AS total
                FROM inspections i
                LEFT JOIN employees cr ON i.created_by = cr.id
                LEFT JOIN employees hd ON i.head_id    = hd.id
                {where_sql}
                ORDER BY i.inspection_date DESC""",
            tuple(params)
        ).fetchall()

    if has_inactive:
        rows = [r for r in rows if (r['inactive'] or 0) > 0]

    return render_template(
        'inspections/history.html',
        rows=rows,
        from_date=from_date,
        to_date=to_date,
        has_inactive=has_inactive,
        email_ready=is_email_configured(),
        is_admin=_is_admin(),
        today_month=date.today().strftime('%Y-%m'),
    )


# ── Excel exports (per-day report + monthly grid) ───────────────────────────

_XLSX_MIME = ('application/vnd.openxmlformats-'
              'officedocument.spreadsheetml.sheet')

# Status -> (short label, fill hex, font hex). Light Excel-friendly colours.
_STATUS_XLSX = {
    'active':   ('Active',   'C6EFCE', '006100'),
    'inactive': ('Inactive', 'FFC7CE', '9C0006'),
    'none':     ('None',     'D9D9D9', '555555'),
}
_STATUS_SHORT = {'active': 'A', 'inactive': 'I', 'none': 'N'}


def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(), mimetype=_XLSX_MIME,
        headers={'Content-Disposition': f'attachment; filename="{filename}"'})


def _day_workbook(conn, inspection_date):
    """Build the per-day report workbook. Returns (wb, filename) or (None, None)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    insp_row = conn.execute(
        "SELECT * FROM inspections WHERE inspection_date = ?",
        (inspection_date,)).fetchone()
    if not insp_row:
        return None, None
    view_date = datetime.strptime(inspection_date, '%Y-%m-%d').date()
    inspection = _inspection_with_signatures(conn, insp_row['id'])
    catalog = _load_catalog(conn)
    results = _load_results(conn, insp_row['id'])
    sheets = _build_sheets(conn, catalog, insp_row['id'], results)
    emp = {e['id']: e['name'] for e in conn.execute(
        "SELECT id, name FROM employees").fetchall()}

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inspection'
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    head_fill = PatternFill('solid', fgColor='1F3864')
    head_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = 1
    ws.cell(r, 1, f'SAIL Daily Inspection — {inspection_date} '
                  f'({view_date.strftime("%A")})').font = title_font
    r += 1
    status_txt = ('Submitted ' + (inspection['submitted_at'] or '')
                  if inspection['submitted_at'] else 'Draft (not submitted)')
    ws.cell(r, 1, status_txt); r += 2

    # Per-sheet summary + sign-off.
    ws.cell(r, 1, 'Sheet').font = bold
    ws.cell(r, 2, 'Owner').font = bold
    ws.cell(r, 3, 'Done').font = bold
    ws.cell(r, 4, 'Inactive').font = bold
    ws.cell(r, 5, 'Completed by').font = bold
    r += 1
    for s in sheets:
        ws.cell(r, 1, s['name'])
        ws.cell(r, 2, s['assignee_name'] or '—')
        ws.cell(r, 3, f"{s['recorded']}/{s['total']} ({s['pct']}%)")
        ws.cell(r, 4, s['inactive'])
        so = s['signoff']
        ws.cell(r, 5, (so['signer_name'] or 'signed') if so else '—')
        r += 1
    r += 1

    # Day sign-off (the head).
    ws.cell(r, 1, 'Day sign-off (Head)').font = bold
    ws.cell(r, 2, inspection['head_name'] or '— not signed —'); r += 1
    ws.cell(r, 1, 'Submitted').font = bold
    ws.cell(r, 2, inspection['submitted_at'] or '—'); r += 2

    # Item table, grouped sheet -> section -> area.
    headers = ['Item', 'Status', 'Note', 'Recorded by', 'Updated']
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(r, c, h)
        cell.font = head_font
        cell.fill = head_fill
    header_row = r
    r += 1
    for s in sheets:
        ws.cell(r, 1, f"■ {s['name'].upper()} — {s['assignee_name'] or 'Unassigned'}").font = bold
        r += 1
        for section in s['sections']:
            ws.cell(r, 1, section['name']).font = Font(bold=True, italic=True)
            r += 1
            for grp in section['groups']:
                ws.cell(r, 1, '   ' + grp['area']['name']).font = bold
                r += 1
                for it in grp['items']:
                    res = results.get(it['id'])
                    if it.get('is_applicable', 1) == 0:
                        ws.cell(r, 1, '      ' + it['name'])
                        c2 = ws.cell(r, 2, 'N/A')
                        c2.fill = PatternFill('solid', fgColor='F2F2F2')
                        c2.font = Font(italic=True, color='888888')
                    else:
                        ws.cell(r, 1, '      ' + it['name'])
                        st = res['status'] if res else None
                        if st in _STATUS_XLSX:
                            label, fill, fcol = _STATUS_XLSX[st]
                            c2 = ws.cell(r, 2, label)
                            c2.fill = PatternFill('solid', fgColor=fill)
                            c2.font = Font(color=fcol, bold=True)
                        else:
                            ws.cell(r, 2, '—')
                        ws.cell(r, 3, (res['notes'] if res else '') or '')
                        ws.cell(r, 4, emp.get(res['updated_by'], '') if res else '')
                        ws.cell(r, 5, (res['updated_at'] if res else '') or '')
                    for c in range(1, 6):
                        ws.cell(r, c).border = border
                    r += 1

    ws.freeze_panes = ws.cell(header_row + 1, 1)
    widths = [42, 12, 34, 18, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    return wb, f'sail-inspection-{inspection_date}.xlsx'


def _month_workbook(conn, month_str):
    """Build the monthly grid workbook. Returns (wb, filename) or (None, None)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    try:
        year, mon = (int(x) for x in month_str.split('-'))
        first = date(year, mon, 1)
    except (ValueError, TypeError):
        return None, None
    ndays = calendar.monthrange(year, mon)[1]
    last = date(year, mon, ndays)

    catalog = _load_catalog(conn)
    sheets = _build_sheets(conn, catalog, None, {})
    rows = conn.execute(
            """SELECT r.item_id,
                      CAST(strftime('%d', i.inspection_date) AS INTEGER) AS day,
                      r.status
               FROM inspection_results r
               JOIN inspections i      ON i.id = r.inspection_id
               JOIN inspection_items it ON it.id = r.item_id
               WHERE i.inspection_date BETWEEN ? AND ?
                 AND it.is_active = 1 AND it.is_applicable = 1""",
            (first.isoformat(), last.isoformat())).fetchall()
    by_cell = {(row['item_id'], row['day']): row['status'] for row in rows}

    wb = Workbook()
    ws = wb.active
    ws.title = first.strftime('%Y-%m')
    bold = Font(bold=True)
    head_fill = PatternFill('solid', fgColor='1F3864')
    head_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')

    ws.cell(1, 1, f'SAIL Inspection — {first.strftime("%B %Y")}').font = \
        Font(bold=True, size=14)
    hr = 2
    hc = ws.cell(hr, 1, 'Item'); hc.font = head_font; hc.fill = head_fill
    for d in range(1, ndays + 1):
        wk = date(year, mon, d).strftime('%a')[0]   # weekday initial
        c = ws.cell(hr, 1 + d, d)
        c.font = head_font; c.fill = head_fill; c.alignment = center
        c2 = ws.cell(hr + 1, 1 + d, wk)
        c2.font = Font(size=8, color='FFFFFF'); c2.fill = head_fill
        c2.alignment = center
    ws.cell(hr + 1, 1, 'Legend: A=Active  I=Inactive  N=None  (blank)=not recorded').font = \
        Font(italic=True, size=9)
    rr = hr + 2

    fills = {k: PatternFill('solid', fgColor=v[1]) for k, v in _STATUS_XLSX.items()}
    for s in sheets:
        ws.cell(rr, 1, f"■ {s['name'].upper()} — {s['assignee_name'] or 'Unassigned'}").font = bold
        rr += 1
        for section in s['sections']:
            ws.cell(rr, 1, section['name']).font = Font(bold=True, italic=True)
            rr += 1
            for grp in section['groups']:
                ws.cell(rr, 1, '  ' + grp['area']['name']).font = bold
                rr += 1
                for it in grp['items']:
                    applicable = it.get('is_applicable', 1) == 1
                    ws.cell(rr, 1, '    ' + it['name'])
                    for d in range(1, ndays + 1):
                        cell = ws.cell(rr, 1 + d)
                        cell.alignment = center
                        if not applicable:
                            cell.value = 'X'
                            cell.fill = PatternFill('solid', fgColor='F2F2F2')
                            cell.font = Font(color='AAAAAA')
                            continue
                        st = by_cell.get((it['id'], d))
                        if st:
                            cell.value = _STATUS_SHORT[st]
                            cell.fill = fills[st]
                    rr += 1

    ws.freeze_panes = ws.cell(hr + 2, 2)
    ws.column_dimensions['A'].width = 40
    for d in range(1, ndays + 1):
        ws.column_dimensions[ws.cell(1, 1 + d).column_letter].width = 4

    return wb, f'sail-inspection-{first.strftime("%Y-%m")}.xlsx'


@inspections_bp.route('/<inspection_date>/export.xlsx')
def export_day_xlsx(inspection_date):
    """Download one day's full inspection as a formatted workbook."""
    if not _is_staff():
        abort(403)
    try:
        datetime.strptime(inspection_date, '%Y-%m-%d')
    except ValueError:
        abort(404)
    with get_db() as conn:
        wb, fname = _day_workbook(conn, inspection_date)
    if not wb:
        flash(f'No inspection recorded for {inspection_date}.', 'info')
        return redirect(url_for('inspections.history'))
    return _xlsx_response(wb, fname)


@inspections_bp.route('/export-month.xlsx')
def export_month_xlsx():
    """Download the whole month as a grid (items × days)."""
    if not _is_staff():
        abort(403)
    month = (request.args.get('month') or date.today().strftime('%Y-%m')).strip()
    with get_db() as conn:
        wb, fname = _month_workbook(conn, month)
    if not wb:
        flash('Bad month — use the picker (YYYY-MM).', 'error')
        return redirect(url_for('inspections.history'))
    return _xlsx_response(wb, fname)


# ── email a report (admin) ──────────────────────────────────────────────────

def _parse_recipients(raw):
    """Split a free-text recipients string into a clean, de-duped email list."""
    import re
    out, seen = [], set()
    for p in re.split(r'[,;\s]+', (raw or '').strip()):
        p = p.strip()
        if p and '@' in p and '.' in p.rsplit('@', 1)[-1]:
            if p.lower() not in seen:
                seen.add(p.lower())
                out.append(p)
    return out


def _report_email_html(title, message):
    safe_msg = (message or '').replace('<', '&lt;').replace('>', '&gt;')
    note = (f'<p style="white-space:pre-wrap">{safe_msg}</p>'
            if safe_msg else '')
    return (f'<p>Hello,</p><p>Please find attached the <strong>{title}</strong> '
            f'report from the SAIL system.</p>{note}'
            f'<p style="color:#888;font-size:12px">Sent by '
            f'{g.user["name"]} via SAIL.</p>')


@inspections_bp.route('/<inspection_date>/email', methods=['POST'])
def email_day(inspection_date):
    if not _is_admin():
        abort(403)
    back = (url_for('inspections.today')
            if inspection_date == date.today().isoformat()
            else url_for('inspections.detail', inspection_date=inspection_date))
    recipients = _parse_recipients(request.form.get('recipients'))
    message = (request.form.get('message') or '').strip()[:1000]
    if not recipients:
        flash('Enter at least one valid email address.', 'error')
        return redirect(back)
    with get_db() as conn:
        wb, fname = _day_workbook(conn, inspection_date)
    if not wb:
        flash('No inspection recorded for that day.', 'error')
        return redirect(url_for('inspections.history'))
    buf = io.BytesIO()
    wb.save(buf)
    sent = send_email_with_attachment(
        recipients, f'Daily Inspection {inspection_date}',
        _report_email_html(f'Daily Inspection — {inspection_date}', message),
        fname, buf.getvalue(), _XLSX_MIME)
    if sent:
        flash(f'Report emailed to {", ".join(recipients)}.', 'success')
    else:
        flash('Email is not set up on the server (SAIL_SMTP_PASSWORD missing) — '
              'nothing was sent. Use the download button instead.', 'error')
    return redirect(back)


@inspections_bp.route('/email-month', methods=['POST'])
def email_month():
    if not _is_admin():
        abort(403)
    month = (request.form.get('month') or date.today().strftime('%Y-%m')).strip()
    recipients = _parse_recipients(request.form.get('recipients'))
    message = (request.form.get('message') or '').strip()[:1000]
    if not recipients:
        flash('Enter at least one valid email address.', 'error')
        return redirect(url_for('inspections.history'))
    with get_db() as conn:
        wb, fname = _month_workbook(conn, month)
    if not wb:
        flash('Bad month — use YYYY-MM.', 'error')
        return redirect(url_for('inspections.history'))
    buf = io.BytesIO()
    wb.save(buf)
    sent = send_email_with_attachment(
        recipients, f'Monthly Inspection {month}',
        _report_email_html(f'Monthly Inspection — {month}', message),
        fname, buf.getvalue(), _XLSX_MIME)
    if sent:
        flash(f'Monthly report emailed to {", ".join(recipients)}.', 'success')
    else:
        flash('Email is not set up on the server (SAIL_SMTP_PASSWORD missing) — '
              'nothing was sent. Use the download button instead.', 'error')
    return redirect(url_for('inspections.history'))


# ── admin: areas catalog ───────────────────────────────────────────────────

def _block_non_admin():
    if not _is_admin():
        flash('Admin only.', 'error')
        return redirect(url_for('inspections.dashboard'))
    return None


@inspections_bp.route('/admin/areas')
def admin_areas():
    block = _block_non_admin()
    if block:
        return block
    with get_db() as conn:
        rows = conn.execute(
            """SELECT a.*,
                      (SELECT COUNT(*) FROM inspection_items it
                         WHERE it.area_id = a.id AND it.is_active = 1) AS item_count
               FROM inspection_areas a
               ORDER BY a.is_active DESC, a.display_order, a.name COLLATE NOCASE"""
        ).fetchall()
    return render_template('inspections/admin_areas.html', areas=rows,
                           sections=SECTION_ORDER)


@inspections_bp.route('/admin/areas/add', methods=['POST'])
def admin_areas_add():
    block = _block_non_admin()
    if block:
        return block
    name = (request.form.get('name') or '').strip()
    order = request.form.get('display_order') or 0
    section = (request.form.get('section') or '').strip()
    if not name:
        flash('Name is required.', 'error')
        return redirect(url_for('inspections.admin_areas'))
    try:
        order = int(order)
    except ValueError:
        order = 0
    with get_db() as conn:
        try:
            cur = conn.execute(
                "INSERT INTO inspection_areas (name, display_order, section) "
                "VALUES (?, ?, ?)", (name, order, section))
            log_audit(conn, 'inspection_areas', cur.lastrowid, 'create',
                      changed_by=g.user['id'])
            flash(f'Added area "{name}".', 'success')
        except Exception as e:
            if 'UNIQUE' in str(e):
                flash('An area with that name already exists.', 'error')
            else:
                raise
    return redirect(url_for('inspections.admin_areas'))


@inspections_bp.route('/admin/areas/<int:area_id>/toggle', methods=['POST'])
def admin_areas_toggle(area_id):
    block = _block_non_admin()
    if block:
        return block
    with get_db() as conn:
        row = conn.execute(
            "SELECT name, is_active FROM inspection_areas WHERE id = ?",
            (area_id,)).fetchone()
        if not row:
            flash('Area not found.', 'error')
            return redirect(url_for('inspections.admin_areas'))
        new_active = 0 if row['is_active'] else 1
        conn.execute(
            "UPDATE inspection_areas SET is_active = ? WHERE id = ?",
            (new_active, area_id))
        log_audit(conn, 'inspection_areas', area_id, 'status_change',
                  'is_active', row['is_active'], new_active,
                  changed_by=g.user['id'])
        flash(f'"{row["name"]}" '
              f'{"reactivated" if new_active else "deactivated"}.', 'success')
    return redirect(url_for('inspections.admin_areas'))


# ── admin: items catalog ───────────────────────────────────────────────────

@inspections_bp.route('/admin/items')
def admin_items():
    block = _block_non_admin()
    if block:
        return block
    with get_db() as conn:
        areas = conn.execute(
            "SELECT id, name, display_order FROM inspection_areas "
            "WHERE is_active = 1 ORDER BY display_order, name"
        ).fetchall()
        items = conn.execute(
            """SELECT it.*, ar.name AS area_name
               FROM inspection_items it
               JOIN inspection_areas ar ON it.area_id = ar.id
               ORDER BY ar.display_order, ar.name, it.display_order, it.name"""
        ).fetchall()
    by_area = {}
    for it in items:
        by_area.setdefault(it['area_id'], []).append(dict(it))
    return render_template(
        'inspections/admin_items.html',
        areas=areas,
        items_by_area=by_area,
    )


@inspections_bp.route('/admin/items/add', methods=['POST'])
def admin_items_add():
    block = _block_non_admin()
    if block:
        return block
    try:
        area_id = int(request.form.get('area_id') or 0)
    except ValueError:
        area_id = 0
    name = (request.form.get('name') or '').strip()
    order_raw = request.form.get('display_order') or 0
    try:
        order = int(order_raw)
    except ValueError:
        order = 0
    if not area_id or not name:
        flash('Area and name are both required.', 'error')
        return redirect(url_for('inspections.admin_items'))
    with get_db() as conn:
        try:
            cur = conn.execute(
                "INSERT INTO inspection_items (area_id, name, display_order) "
                "VALUES (?, ?, ?)", (area_id, name, order))
            log_audit(conn, 'inspection_items', cur.lastrowid, 'create',
                      changed_by=g.user['id'])
            flash(f'Added "{name}".', 'success')
        except Exception as e:
            if 'UNIQUE' in str(e):
                flash('That item already exists in this area.', 'error')
            else:
                raise
    return redirect(url_for('inspections.admin_items'))


@inspections_bp.route('/admin/items/<int:item_id>/toggle', methods=['POST'])
def admin_items_toggle(item_id):
    block = _block_non_admin()
    if block:
        return block
    with get_db() as conn:
        row = conn.execute(
            "SELECT name, is_active FROM inspection_items WHERE id = ?",
            (item_id,)).fetchone()
        if not row:
            flash('Item not found.', 'error')
            return redirect(url_for('inspections.admin_items'))
        new_active = 0 if row['is_active'] else 1
        conn.execute(
            "UPDATE inspection_items SET is_active = ? WHERE id = ?",
            (new_active, item_id))
        log_audit(conn, 'inspection_items', item_id, 'status_change',
                  'is_active', row['is_active'], new_active,
                  changed_by=g.user['id'])
        flash(f'"{row["name"]}" '
              f'{"reactivated" if new_active else "deactivated"}.', 'success')
    return redirect(url_for('inspections.admin_items'))


@inspections_bp.route('/admin/items/<int:item_id>/toggle-na', methods=['POST'])
def admin_items_toggle_na(item_id):
    """Flip an item between applicable and N/A (not applicable for this room)."""
    block = _block_non_admin()
    if block:
        return block
    with get_db() as conn:
        row = conn.execute(
            "SELECT name, is_applicable FROM inspection_items WHERE id = ?",
            (item_id,)).fetchone()
        if not row:
            flash('Item not found.', 'error')
            return redirect(url_for('inspections.admin_items'))
        new_applicable = 0 if row['is_applicable'] else 1
        conn.execute(
            "UPDATE inspection_items SET is_applicable = ? WHERE id = ?",
            (new_applicable, item_id))
        log_audit(conn, 'inspection_items', item_id, 'status_change',
                  'is_applicable', row['is_applicable'], new_applicable,
                  changed_by=g.user['id'])
        flash(f'"{row["name"]}" marked '
              f'{"applicable" if new_applicable else "N/A"}.', 'success')
    return redirect(url_for('inspections.admin_items'))


# ── admin: sheets & assignees ───────────────────────────────────────────────

@inspections_bp.route('/admin/sheets')
def admin_sheets():
    block = _block_non_admin()
    if block:
        return block
    with get_db() as conn:
        sheets = conn.execute(
            """SELECT s.name, s.display_order, s.assignee_id, e.name AS assignee_name
               FROM inspection_sheets s
               LEFT JOIN employees e ON e.id = s.assignee_id
               ORDER BY s.display_order, s.name""").fetchall()
        employees = conn.execute(
            "SELECT id, name, role FROM employees "
            "WHERE is_active = 1 ORDER BY name").fetchall()
    # Which sections belong to each sheet (for display).
    sheet_sections = {}
    for section, sheet in SHEET_OF_SECTION.items():
        sheet_sections.setdefault(sheet, []).append(section)
    return render_template('inspections/admin_sheets.html',
                           sheets=sheets, employees=employees,
                           sheet_sections=sheet_sections)


@inspections_bp.route('/admin/sheets/assign', methods=['POST'])
def admin_sheets_assign():
    block = _block_non_admin()
    if block:
        return block
    sheet = (request.form.get('sheet') or '').strip()
    raw = (request.form.get('assignee_id') or '').strip()
    assignee_id = int(raw) if raw.isdigit() else None
    with get_db() as conn:
        row = conn.execute(
            "SELECT name, assignee_id FROM inspection_sheets WHERE name = ?",
            (sheet,)).fetchone()
        if not row:
            flash('Unknown sheet.', 'error')
            return redirect(url_for('inspections.admin_sheets'))
        conn.execute(
            "UPDATE inspection_sheets SET assignee_id = ? WHERE name = ?",
            (assignee_id, sheet))
        log_audit(conn, 'inspection_sheets', 0, 'status_change',
                  f'assignee:{sheet}', row['assignee_id'], assignee_id,
                  changed_by=g.user['id'])
        flash(f'{sheet} assignee updated.', 'success')
    return redirect(url_for('inspections.admin_sheets'))
