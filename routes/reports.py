"""Reports — weekly/monthly history + current-status rollups for inventory and tickets."""
import csv
import io
from datetime import date, datetime, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, g, flash, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from database import get_db

reports_bp = Blueprint('reports', __name__)

PERIODS = {'week': 7, 'month': 30}


def _admin_only():
    if not g.user or g.user['role'] not in ('admin', 'manager', 'technician'):
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard.index'))
    return None


def _resolve_range():
    """Return (period, start_iso, end_iso, end_date) from request args."""
    period = request.args.get('period', 'month')
    if period not in PERIODS:
        period = 'month'

    end_str = request.args.get('end', '').strip()
    try:
        end_d = datetime.strptime(end_str, '%Y-%m-%d').date() if end_str else date.today()
    except ValueError:
        end_d = date.today()

    start_d = end_d - timedelta(days=PERIODS[period] - 1)
    # end boundary is exclusive upper (next day 00:00) so created_at comparisons catch the whole day
    return period, start_d.isoformat(), (end_d + timedelta(days=1)).isoformat(), end_d


def _csv_response(filename, headers, rows):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
    return Response(
        buf.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'})


# ── Inventory report ──────────────────────────────────────────────────

@reports_bp.route('/inventory')
def inventory():
    guard = _admin_only()
    if guard:
        return guard

    period, start, end, end_d = _resolve_range()
    fmt = request.args.get('format', '')

    with get_db() as conn:
        # Current snapshot
        status_breakdown = conn.execute("""
            SELECT status, COUNT(*) as n FROM assets GROUP BY status ORDER BY n DESC
        """).fetchall()
        condition_breakdown = conn.execute("""
            SELECT condition, COUNT(*) as n FROM assets GROUP BY condition ORDER BY n DESC
        """).fetchall()
        category_breakdown = conn.execute("""
            SELECT c.name as category, COUNT(a.id) as n
            FROM categories c
            LEFT JOIN equipment_models em ON em.category_id = c.id
            LEFT JOIN assets a ON a.equipment_model_id = em.id
            GROUP BY c.name ORDER BY n DESC
        """).fetchall()

        totals = {
            'assets': conn.execute("SELECT COUNT(*) FROM assets").fetchone()[0],
            'models': conn.execute("SELECT COUNT(*) FROM equipment_models").fetchone()[0],
            'available': conn.execute(
                "SELECT COUNT(*) FROM assets WHERE status='available'").fetchone()[0],
            'checked_out': conn.execute(
                "SELECT COUNT(*) FROM assets WHERE status='checked_out'").fetchone()[0],
            'maintenance': conn.execute(
                "SELECT COUNT(*) FROM assets WHERE status='maintenance'").fetchone()[0],
        }

        # History within period
        history = {
            'new_assets': conn.execute(
                "SELECT COUNT(*) FROM assets WHERE created_at >= ? AND created_at < ?",
                (start, end)).fetchone()[0],
        }

    if fmt == 'csv':
        rows = [
            ['SECTION', 'KEY', 'VALUE'],
            ['Period', 'period', period],
            ['Period', 'start', start],
            ['Period', 'end', end_d.isoformat()],
        ]
        for k, v in totals.items():
            rows.append(['Current', k, v])
        for k, v in history.items():
            rows.append(['History', k, v])
        for r in status_breakdown:
            rows.append(['Status', r['status'], r['n']])
        for r in condition_breakdown:
            rows.append(['Condition', r['condition'], r['n']])
        for r in category_breakdown:
            rows.append(['Category', r['category'], r['n']])
        fname = f"inventory-report-{period}-{end_d.isoformat()}.csv"
        return _csv_response(fname, rows[0], rows[1:])

    return render_template('reports/inventory.html',
                           period=period, start=start, end_d=end_d,
                           totals=totals, history=history,
                           status_breakdown=status_breakdown,
                           condition_breakdown=condition_breakdown,
                           category_breakdown=category_breakdown)


# ── Ticket report ─────────────────────────────────────────────────────

@reports_bp.route('/tickets')
def tickets():
    guard = _admin_only()
    if guard:
        return guard

    period, start, end, end_d = _resolve_range()
    fmt = request.args.get('format', '')

    with get_db() as conn:
        # Current snapshot (all-time open/in-progress etc.)
        status_breakdown = conn.execute("""
            SELECT status, COUNT(*) as n FROM tickets GROUP BY status ORDER BY n DESC
        """).fetchall()
        priority_breakdown = conn.execute("""
            SELECT priority, COUNT(*) as n FROM tickets
            WHERE status IN ('open','in_progress','waiting')
            GROUP BY priority
        """).fetchall()
        type_breakdown = conn.execute("""
            SELECT type, COUNT(*) as n FROM tickets GROUP BY type ORDER BY n DESC
        """).fetchall()

        totals = {
            'all_tickets': conn.execute("SELECT COUNT(*) FROM tickets").fetchone()[0],
            'open': conn.execute(
                "SELECT COUNT(*) FROM tickets WHERE status IN ('open','in_progress','waiting')"
            ).fetchone()[0],
            'unassigned_open': conn.execute(
                "SELECT COUNT(*) FROM tickets WHERE assigned_to IS NULL "
                "AND status IN ('open','in_progress','waiting')"
            ).fetchone()[0],
            'critical_open': conn.execute(
                "SELECT COUNT(*) FROM tickets WHERE priority='critical' "
                "AND status IN ('open','in_progress','waiting')"
            ).fetchone()[0],
        }

        # History within period
        history = {
            'created': conn.execute(
                "SELECT COUNT(*) FROM tickets WHERE created_at >= ? AND created_at < ?",
                (start, end)).fetchone()[0],
            'resolved': conn.execute(
                "SELECT COUNT(*) FROM tickets "
                "WHERE resolved_at IS NOT NULL AND resolved_at >= ? AND resolved_at < ?",
                (start, end)).fetchone()[0],
            'closed': conn.execute(
                "SELECT COUNT(*) FROM tickets "
                "WHERE closed_at IS NOT NULL AND closed_at >= ? AND closed_at < ?",
                (start, end)).fetchone()[0],
        }

        # Avg resolution time for tickets resolved in the period (hours)
        avg_row = conn.execute("""
            SELECT AVG((julianday(resolved_at) - julianday(created_at)) * 24.0) as avg_h
            FROM tickets
            WHERE resolved_at IS NOT NULL AND resolved_at >= ? AND resolved_at < ?
        """, (start, end)).fetchone()
        history['avg_resolution_hours'] = round(avg_row['avg_h'], 1) if avg_row['avg_h'] else None

        created_by_type = conn.execute("""
            SELECT type, COUNT(*) as n FROM tickets
            WHERE created_at >= ? AND created_at < ?
            GROUP BY type ORDER BY n DESC
        """, (start, end)).fetchall()

        created_by_priority = conn.execute("""
            SELECT priority, COUNT(*) as n FROM tickets
            WHERE created_at >= ? AND created_at < ?
            GROUP BY priority
        """, (start, end)).fetchall()

        top_assignees = conn.execute("""
            SELECT e.name as assignee,
                   SUM(CASE WHEN t.resolved_at >= ? AND t.resolved_at < ? THEN 1 ELSE 0 END) as resolved,
                   SUM(CASE WHEN t.status IN ('open','in_progress','waiting') THEN 1 ELSE 0 END) as open_now
            FROM tickets t
            JOIN employees e ON t.assigned_to = e.id
            WHERE t.assigned_to IS NOT NULL
            GROUP BY e.id
            HAVING resolved > 0 OR open_now > 0
            ORDER BY resolved DESC, open_now DESC LIMIT 10
        """, (start, end)).fetchall()

        recent_tickets = conn.execute("""
            SELECT t.id, t.ticket_number, t.title, t.type, t.priority, t.status,
                   t.created_at, t.resolved_at,
                   sub.name as submitter, asg.name as assignee
            FROM tickets t
            JOIN employees sub ON t.submitted_by = sub.id
            LEFT JOIN employees asg ON t.assigned_to = asg.id
            WHERE t.created_at >= ? AND t.created_at < ?
            ORDER BY t.created_at DESC LIMIT 50
        """, (start, end)).fetchall()

        # Backlog — open tickets older than the period start
        aging_backlog = conn.execute("""
            SELECT t.ticket_number, t.title, t.type, t.priority, t.created_at,
                   sub.name as submitter, asg.name as assignee,
                   CAST(julianday('now') - julianday(t.created_at) AS INTEGER) as age_days
            FROM tickets t
            JOIN employees sub ON t.submitted_by = sub.id
            LEFT JOIN employees asg ON t.assigned_to = asg.id
            WHERE t.status IN ('open','in_progress','waiting')
              AND t.created_at < ?
            ORDER BY t.created_at ASC LIMIT 20
        """, (start,)).fetchall()

    if fmt == 'csv':
        rows = [
            ['SECTION', 'KEY', 'VALUE'],
            ['Period', 'period', period],
            ['Period', 'start', start],
            ['Period', 'end', end_d.isoformat()],
        ]
        for k, v in totals.items():
            rows.append(['Current', k, v])
        for k, v in history.items():
            rows.append(['History', k, v if v is not None else ''])
        for r in status_breakdown:
            rows.append(['Status', r['status'], r['n']])
        for r in type_breakdown:
            rows.append(['Type', r['type'], r['n']])
        for r in priority_breakdown:
            rows.append(['OpenPriority', r['priority'], r['n']])
        for r in created_by_type:
            rows.append(['CreatedByType', r['type'], r['n']])
        for r in top_assignees:
            rows.append(['Assignee', r['assignee'], f"resolved={r['resolved']}, open={r['open_now']}"])
        fname = f"tickets-report-{period}-{end_d.isoformat()}.csv"
        return _csv_response(fname, rows[0], rows[1:])

    return render_template('reports/tickets.html',
                           period=period, start=start, end_d=end_d,
                           totals=totals, history=history,
                           status_breakdown=status_breakdown,
                           priority_breakdown=priority_breakdown,
                           type_breakdown=type_breakdown,
                           created_by_type=created_by_type,
                           created_by_priority=created_by_priority,
                           top_assignees=top_assignees,
                           recent_tickets=recent_tickets,
                           aging_backlog=aging_backlog)


# ── Full database dump to Excel ───────────────────────────────────────

# Order matters — lookup/reference tables first so cross-sheet navigation is intuitive.
EXPORT_TABLES = [
    'categories', 'locations', 'departments', 'employees',
    'equipment_models', 'assets',
    'tickets', 'ticket_comments', 'audit_log',
]

HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2D3748')


@reports_bp.route('/export-all.xlsx')
def export_all_xlsx():
    """One workbook, one sheet per table — full data dump for admins."""
    guard = _admin_only()
    if guard:
        return guard

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    with get_db() as conn:
        for table in EXPORT_TABLES:
            rows = conn.execute(f"SELECT * FROM {table}").fetchall()
            ws = wb.create_sheet(title=table[:31])  # Excel sheet-name limit

            if not rows:
                ws.append(['(no rows)'])
                continue

            headers = list(rows[0].keys())
            ws.append(headers)
            for cell in ws[1]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL

            for row in rows:
                ws.append([row[h] for h in headers])

            ws.freeze_panes = 'A2'
            for i, h in enumerate(headers, start=1):
                sample = str(rows[0][h] or '')
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = \
                    min(max(len(h), len(sample), 10), 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"sail-full-export-{date.today().isoformat()}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'})
