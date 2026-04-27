"""
Import the V3 asset inventory into SAIL.

Usage:
    python import_assets_v3.py              # full import (wipes + reloads)
    python import_assets_v3.py --dry-run    # parse + summarize, no DB writes
    python import_assets_v3.py --xlsx PATH  # override source Excel path

The Excel sheet "IT Assets" is the source of truth. Categories,
locations, and equipment_models are derived from the row data.
See docs/superpowers/specs/2026-04-27-asset-data-bootstrap-design.md.
"""
import argparse
import os
import re
import sys
from collections import Counter

import openpyxl

from database import get_db
from backup_db import backup as backup_db

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(
    BASE_DIR, "Assets Inventory _20-04-2026-Tool (V3).xlsx"
)
SHEET_NAME = "IT Assets"

# Category names that should NOT default to bookable=1 (fixed installations).
NON_BOOKABLE_CATEGORIES = {"Access Control", "Smart Podium", "Eye Tracking System"}

# Holder values that mean "in the SAIL inventory pool" (status = available).
STORAGE_POOL_HOLDERS = {"SAIL", "SAIL Storage", "-"}

# Header → row-index map (filled by read_rows).
EXPECTED_HEADERS = {
    "sequence": "Sequence",
    "product_id": "Product_ID(SAIL ID)",
    "category": "Category",
    "item_name": "Item Name",
    "description": "Description",
    "availability": "Availability",
    "holder_name": "Holder Name",
    "serial_number": "Serial Number",
    "desk_area": "Desk/Site Area",
    "official_location": "Official location",
    "remark": "Remark",
    "date_from": "Date From",
    "date_to": "Date To",
    "image": "Image",
    "phone": "phone",
    "email": "Email",
}


# ── Pure transformation helpers ─────────────────────────────────────────────


def s(v):
    """Coerce a cell to a stripped string, or None for empty/whitespace."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return str(v).strip() or None


def normalize_category(raw):
    """Fold "MONITOR", "Smart board", etc. to a canonical title-case form."""
    if not raw:
        return None
    return raw.strip().title()


def normalize_item_name(raw):
    """Item name used as a model-grouping key. Just trims; preserves brand casing."""
    return raw.strip() if raw else None


def derive_asset_tag(product_id, row_num, dup_pids):
    """
    Per spec §5.2:
      - unique PID         -> SAIL-{pid}
      - duplicated PID     -> SAIL-{pid}-R{row}    (PID + Excel row for traceability)
      - missing PID        -> SAIL-ROW-{row}
    `dup_pids` is the set of PID values that appear on more than one row in the
    sheet; the caller is responsible for computing it in a first pass.
    """
    if not product_id:
        return f"SAIL-ROW-{row_num}"
    if product_id in dup_pids:
        return f"SAIL-{product_id}-R{row_num}"
    return f"SAIL-{product_id}"


def derive_condition(availability):
    """Excel Availability → assets.condition."""
    if not availability:
        return "good"
    a = str(availability).strip().lower()
    if a in ("yes", "1"):
        return "good"
    if a == "damage":
        return "damaged"
    if a == "no":
        return "fair"
    return "good"  # unknown → assume good; let admins fix


def derive_status(holder, remark):
    """
    Derived per the §6 rules in the spec:
      - Remark = Not Found/Missing  -> missing
      - Holder = NOT SAIL           -> decommissioned
      - Holder in storage-pool      -> available
      - Otherwise                   -> in_use
    The 'Found Not in App' remark falls through to the holder rules.
    """
    if remark and remark.strip().lower() == "not found/missing":
        return "missing"
    h = (holder or "").strip()
    if h.upper() == "NOT SAIL":
        return "decommissioned"
    if not h or h in STORAGE_POOL_HOLDERS:
        return "available"
    return "in_use"


def location_code(label):
    """Slug for locations.code: uppercase, slashes/spaces -> '-', collapse runs."""
    if not label:
        return "UNKNOWN"
    code = label.strip().upper()
    code = re.sub(r"[\s/]+", "-", code)
    code = re.sub(r"-+", "-", code).strip("-")
    return code or "UNKNOWN"


def location_for(raw_label):
    """
    Map an "Official location" cell to (code, label, is_storage).
    Blank or 'N/A' collapse to the single UNKNOWN location.
    """
    if not raw_label or raw_label.strip().upper() in ("N/A", ""):
        return ("UNKNOWN", "Unknown / N-A", 0)
    label = raw_label.strip()
    code = location_code(label)
    is_storage = 1 if label.upper() == "STORAGE" else 0
    return (code, label, is_storage)


def is_bookable_for(category):
    """Default bookability flag for a derived equipment_model."""
    return 0 if category in NON_BOOKABLE_CATEGORIES else 1


def build_notes(row):
    """Fold sparse Excel fields into a single notes string for an asset."""
    parts = []
    desk = s(row.get("desk_area"))
    off = s(row.get("official_location"))
    if desk and (not off or desk.lower() != off.lower()):
        parts.append(f"desk: {desk}")
    for key in ("date_from", "date_to", "phone", "email"):
        val = s(row.get(key))
        if val:
            parts.append(f"{key}: {val}")
    return " | ".join(parts) if parts else None


# ── Excel reader ────────────────────────────────────────────────────────────


def read_rows(xlsx_path):
    """Yield (excel_row_num, dict-keyed-by-EXPECTED_HEADERS) per data row.

    excel_row_num is the 1-based row number in the source sheet (the header
    is row 1; the first data row is row 2). It is used by derive_asset_tag
    as the deterministic uniqueness fallback (see spec §5.2).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"sheet {SHEET_NAME!r} not found in {xlsx_path}")
    ws = wb[SHEET_NAME]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = {}
    for key, label in EXPECTED_HEADERS.items():
        try:
            idx[key] = header_row.index(label)
        except ValueError:
            raise SystemExit(
                f"missing expected header {label!r} in {xlsx_path} (row 1)"
            )
    for row_num, raw in enumerate(ws.iter_rows(min_row=2, values_only=True),
                                  start=2):
        # Skip totally empty rows (no Sequence AND no Category).
        if raw[idx["sequence"]] is None and raw[idx["category"]] is None:
            continue
        yield row_num, {key: raw[i] for key, i in idx.items()}


# ── Derivation pass (no DB) ─────────────────────────────────────────────────


def derive_all(rows):
    """
    Walk the rows (each is (row_num, dict)) and build the in-memory plan:
      categories: set of names
      locations:  dict {code: (label, is_storage)}
      models:     dict {(category, item_name_lower): {category, name, image, specs}}
      assets:     list of asset-row dicts ready for INSERT

    The walk is two-pass — the first pass identifies duplicate Product_IDs so
    derive_asset_tag can suffix them with the Excel row number (spec §5.2).
    """
    rows = list(rows)  # materialise so we can iterate twice

    # Pass 1: count Product_IDs to find duplicates.
    pid_counter = Counter()
    for _row_num, row in rows:
        pid = s(row["product_id"])
        if pid:
            pid_counter[pid] += 1
    dup_pids = {pid for pid, n in pid_counter.items() if n > 1}

    categories = set()
    locations = {}
    models = {}
    assets = []

    status_counter = Counter()
    no_pid = 0
    dup_pid_rows = 0
    badge_holders = 0
    found_not_in_app = 0

    # Pass 2: build the plan.
    for row_num, row in rows:
        cat = normalize_category(s(row["category"]))
        item = normalize_item_name(s(row["item_name"]))
        if not cat:
            continue  # safety; read_rows already filters totally empty rows

        categories.add(cat)

        loc_code, loc_label, loc_is_storage = location_for(s(row["official_location"]))
        if loc_code not in locations:
            locations[loc_code] = (loc_label, loc_is_storage)

        item_key = (cat, (item or "").lower())
        if item_key not in models:
            models[item_key] = {
                "category": cat,
                "name": item or cat,
                "description": None,
                "image": None,
            }
        m = models[item_key]
        if m["description"] is None:
            desc = s(row["description"])
            if desc and desc.lower() != (item or "").lower():
                m["description"] = desc
        if m["image"] is None:
            img = s(row["image"])
            if img and img.lower().startswith("http"):
                m["image"] = img

        product_id = s(row["product_id"])
        if not product_id:
            no_pid += 1
        elif product_id in dup_pids:
            dup_pid_rows += 1
        asset_tag = derive_asset_tag(product_id, row_num, dup_pids)

        holder = s(row["holder_name"])
        remark = s(row["remark"])
        status = derive_status(holder, remark)
        condition = derive_condition(s(row["availability"]))

        if remark and remark.lower() == "found not in app":
            found_not_in_app += 1
        if holder and re.search(r"\d{5,}", holder):
            badge_holders += 1

        status_counter[status] += 1

        assets.append({
            "asset_tag": asset_tag,
            "model_key": item_key,
            "loc_code": loc_code,
            "category": cat,
            "serial_number": s(row["serial_number"]),
            "condition": condition,
            "status": status,
            "holder_name": holder,
            "remark": remark,
            "image_path": m["image"],   # mirrored on the model; not used per-asset
            "notes": build_notes(row),
        })

    # Sanity: tags must be unique by construction. If they aren't, something
    # in derive_asset_tag is broken.
    tag_counter = Counter(a["asset_tag"] for a in assets)
    duplicates = [t for t, n in tag_counter.items() if n > 1]
    if duplicates:
        raise RuntimeError(f"asset_tag uniqueness violation: {duplicates[:5]} ...")

    return {
        "categories": categories,
        "locations": locations,
        "models": models,
        "assets": assets,
        "status_counter": status_counter,
        "no_pid": no_pid,
        "dup_pid_rows": dup_pid_rows,
        "badge_holders": badge_holders,
        "found_not_in_app": found_not_in_app,
    }


# ── Output ──────────────────────────────────────────────────────────────────


def print_summary(plan):
    cats = plan["categories"]
    locs = plan["locations"]
    models = plan["models"]
    assets = plan["assets"]
    sc = plan["status_counter"]
    bookable = sum(
        1 for (cat, _) in models.keys() if is_bookable_for(cat)
    )

    print("SUMMARY")
    print(f"  Categories:        {len(cats)}")
    print(f"  Locations:         {len(locs)}")
    print(f"  Equipment models:  {len(models)}")
    print(f"  Assets:            {len(assets)}")
    for st in ("available", "in_use", "missing", "decommissioned",
               "reserved", "checked_out", "maintenance"):
        if sc.get(st):
            print(f"    {st + ':':<16} {sc[st]}")
    print(f"  Bookable models:   {bookable} of {len(models)}")
    print("DATA QUALITY")
    print(f"  Rows w/o Product_ID:        {plan['no_pid']}   (assigned SAIL-ROW-{{excel_row}})")
    print(f"  Rows w/ duplicate PID:      {plan['dup_pid_rows']}   (suffixed with -R{{excel_row}})")
    print(f"  Rows w/ holder badge#:      {plan['badge_holders']}")
    print(f"  Rows w/ \"Found Not in App\": {plan['found_not_in_app']}")


# ── Entry point (DB write path stubbed for now) ─────────────────────────────


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true",
                        help="parse + summarize, no DB writes")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX,
                        help=f"path to the V3 Excel (default: {DEFAULT_XLSX})")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        sys.exit(f"Excel not found: {args.xlsx}")

    rows = list(read_rows(args.xlsx))   # list of (row_num, row_dict) tuples
    print(f"Read {len(rows)} rows from {args.xlsx}")

    plan = derive_all(rows)
    print_summary(plan)

    if args.dry_run:
        print("\n--dry-run: no DB writes performed")
        return

    # Task 4 wires up the actual DB write.
    sys.exit("DB write path not yet implemented — re-run with --dry-run for now")


if __name__ == "__main__":
    main()
